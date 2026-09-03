"""
Professional Report Exporter (Python 3.12+).

Exports to:
  - Excel (.xlsx): styled headers, zebra rows, conditional formatting,
                   Executive Summary sheet
  - CSV   (.csv) : UTF-8 with BOM for Excel compatibility
  - JSON  (.json): machine-readable, pretty-printed
  - HTML  (index.html): injects reportData array into live dashboard

10 report columns:
  Website | URL | Status | CSS Framework(s) | Design System(s) Used |
  UX4G Design System (Yes/No) | UX4G Accessibility CDN (Yes/No) |
  Confidence | Evidence | Notes
"""

from __future__ import annotations

import json
import os
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

HEADERS = [
    "Website",
    "URL",
    "Status",
    "CSS Framework(s)",
    "Design System(s) Used",
    "UX4G Design System (Yes/No)",
    "UX4G Accessibility CDN (Yes/No)",
    "Confidence",
    "Evidence",
    "Notes",
]

COL_WIDTHS = {
    1: 28,   # Website
    2: 38,   # URL
    3: 12,   # Status
    4: 28,   # CSS Framework(s)
    5: 34,   # Design System(s) Used
    6: 20,   # UX4G Design System
    7: 24,   # UX4G Accessibility CDN
    8: 13,   # Confidence
    9: 65,   # Evidence
    10: 32,  # Notes
}

# Colour palette
C_HEADER_BG   = "1E293B"
C_GREEN_BG    = "DCFCE7"
C_GREEN_TEXT  = "166534"
C_RED_BG      = "FEE2E2"
C_RED_TEXT    = "991B1B"
C_ZEBRA       = "F8FAFC"
C_HIGH        = "1D4ED8"
C_MEDIUM      = "B45309"
C_LOW         = "64748B"
C_LIVE_BG     = "D1FAE5"
C_LIVE_FG     = "065F46"
C_VERIFY_BG   = "FEF3C7"
C_VERIFY_FG   = "92400E"
C_ERROR_BG    = "FEE2E2"
C_ERROR_FG    = "991B1B"
C_SKIP_BG     = "F1F5F9"
C_SKIP_FG     = "64748B"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(
    *,
    bold: bool = False,
    color: str = "0F172A",
    size: int = 10,
    name: str = "Calibri",
) -> Font:
    return Font(name=name, size=size, bold=bold, color=color)


def _sanitize(value: Any) -> str:
    if not isinstance(value, str):
        value = str(value) if value is not None else ""
    return " ".join(value.replace("\r\n", " ").replace("\n", " ").split())


# ─────────────────────────────────────────────────────────────────────────────
# Exporter
# ─────────────────────────────────────────────────────────────────────────────

class ReportExporter:
    def __init__(self, output_dir: str) -> None:
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    # ── CSV ───────────────────────────────────────────────────────────────────

    def export_csv(
        self,
        records: list[dict[str, Any]],
        filename: str = "Design_System_Report.csv",
    ) -> str:
        filepath = os.path.join(self.output_dir, filename)
        df = pd.DataFrame(records, columns=HEADERS)
        df.to_csv(filepath, index=False, encoding="utf-8-sig")
        return filepath

    # ── JSON ──────────────────────────────────────────────────────────────────

    def export_json(
        self,
        records: list[dict[str, Any]],
        filename: str = "Design_System_Report.json",
    ) -> str:
        filepath = os.path.join(self.output_dir, filename)
        with open(filepath, "w", encoding="utf-8") as fh:
            json.dump(records, fh, indent=2, ensure_ascii=False)
        return filepath

    # ── Excel ─────────────────────────────────────────────────────────────────

    def export_excel(
        self,
        records: list[dict[str, Any]],
        filename: str = "Design_System_Report.xlsx",
    ) -> str:
        filepath = os.path.join(self.output_dir, filename)
        wb = openpyxl.Workbook()

        self._build_report_sheet(wb, records)
        self._build_summary_sheet(wb, records)

        wb.save(filepath)
        return filepath

    def _build_report_sheet(
        self, wb: openpyxl.Workbook, records: list[dict[str, Any]]
    ) -> None:
        ws = wb.active
        ws.title = "Technology Report"

        border = _thin_border()
        hdr_font  = _font(bold=True, color="FFFFFF", size=11)
        hdr_fill  = _fill(C_HEADER_BG)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── Header row ────────────────────────────────────────────────────────
        ws.append(HEADERS)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font   = hdr_font
            cell.fill   = hdr_fill
            cell.border = border
            cell.alignment = hdr_align
        ws.row_dimensions[1].height = 30

        # ── Data rows ─────────────────────────────────────────────────────────
        zebra = _fill(C_ZEBRA)

        for ri, rec in enumerate(records, start=2):
            row_vals = [_sanitize(rec.get(h, "")) for h in HEADERS]
            ws.append(row_vals)
            ws.row_dimensions[ri].height = 22

            is_zebra = ri % 2 == 0

            for ci in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.border = border
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=(ci in (4, 5, 9, 10)),
                )

                if is_zebra:
                    cell.fill = zebra

                col_header = HEADERS[ci - 1]

                # Status column
                if col_header == "Status":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    val = (cell.value or "").upper()
                    if "LIVE" in val:
                        cell.fill = _fill(C_LIVE_BG)
                        cell.font = _font(bold=True, color=C_LIVE_FG)
                    elif "VERIFY" in val:
                        cell.fill = _fill(C_VERIFY_BG)
                        cell.font = _font(bold=False, color=C_VERIFY_FG)
                    elif "ERROR" in val or "FAIL" in val:
                        cell.fill = _fill(C_ERROR_BG)
                        cell.font = _font(bold=False, color=C_ERROR_FG)
                    elif "SKIP" in val or "NO URL" in val:
                        cell.fill = _fill(C_SKIP_BG)
                        cell.font = _font(color=C_SKIP_FG)

                # UX4G DS column
                elif col_header == "UX4G Design System (Yes/No)":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.value == "Yes":
                        cell.fill = _fill(C_GREEN_BG)
                        cell.font = _font(bold=True, color=C_GREEN_TEXT)
                    else:
                        cell.fill = _fill(C_RED_BG)
                        cell.font = _font(color=C_RED_TEXT)

                # UX4G Acc column
                elif col_header == "UX4G Accessibility CDN (Yes/No)":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.value == "Yes":
                        cell.fill = _fill(C_GREEN_BG)
                        cell.font = _font(bold=True, color=C_GREEN_TEXT)
                    else:
                        cell.fill = _fill(C_RED_BG)
                        cell.font = _font(color=C_RED_TEXT)

                # Confidence column
                elif col_header == "Confidence":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.value == "High":
                        cell.font = _font(bold=True, color=C_HIGH)
                    elif cell.value == "Medium":
                        cell.font = _font(bold=True, color=C_MEDIUM)
                    else:
                        cell.font = _font(color=C_LOW)

        # ── Column widths ─────────────────────────────────────────────────────
        for ci, width in COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(ci)].width = width

        # Freeze header
        ws.freeze_panes = "A2"

    def _build_summary_sheet(
        self, wb: openpyxl.Workbook, records: list[dict[str, Any]]
    ) -> None:
        ws = wb.create_sheet(title="Executive Summary")

        title = ws.cell(row=1, column=1, value="Executive Summary — UX4G Design System Audit")
        title.font = Font(name="Calibri", size=14, bold=True, color="0F172A")

        total = len(records)
        live = sum(1 for r in records if "LIVE" in str(r.get("Status", "")).upper())
        errors = sum(1 for r in records if "ERROR" in str(r.get("Status", "")).upper())
        skipped = sum(1 for r in records if r.get("URL") in ("N/A", None, ""))
        ux4g_ds = sum(1 for r in records if r.get("UX4G Design System (Yes/No)") == "Yes")
        ux4g_acc = sum(1 for r in records if r.get("UX4G Accessibility CDN (Yes/No)") == "Yes")
        high = sum(1 for r in records if r.get("Confidence") == "High")
        medium = sum(1 for r in records if r.get("Confidence") == "Medium")
        bootstrap = sum(1 for r in records if "Bootstrap" in str(r.get("CSS Framework(s)", "")))
        tailwind = sum(1 for r in records if "Tailwind" in str(r.get("CSS Framework(s)", "")))

        def pct(n: int) -> str:
            return f"{n} ({n / total * 100:.1f}%)" if total else "0"

        rows = [
            ("Metric", "Value"),
            ("Total Portals Analysed", total),
            ("LIVE Status", pct(live)),
            ("Errors / Unreachable", pct(errors)),
            ("Skipped (No URL)", pct(skipped)),
            ("", ""),
            ("UX4G Design System Adopted", pct(ux4g_ds)),
            ("UX4G Accessibility CDN Adopted", pct(ux4g_acc)),
            ("", ""),
            ("High Confidence Results", pct(high)),
            ("Medium Confidence Results", pct(medium)),
            ("", ""),
            ("Bootstrap Usage", pct(bootstrap)),
            ("Tailwind CSS Usage", pct(tailwind)),
        ]

        hdr_fill = _fill(C_HEADER_BG)
        hdr_font = _font(bold=True, color="FFFFFF", size=11)

        for r_idx, (metric, value) in enumerate(rows, start=3):
            c1 = ws.cell(row=r_idx, column=1, value=metric)
            c2 = ws.cell(row=r_idx, column=2, value=value)
            if r_idx == 3:
                c1.fill = hdr_fill
                c1.font = hdr_font
                c2.fill = hdr_fill
                c2.font = hdr_font
            elif metric:
                c1.font = _font(bold=False)
                c2.font = _font(bold=True, color=C_HIGH)

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 26

    # ── HTML dashboard injection ──────────────────────────────────────────────

    def update_index_html(
        self, records: list[dict[str, Any]], html_file_path: str
    ) -> None:
        """
        Injects the latest results into the index.html reportData constant.
        Sanitizes all string values before encoding.
        """
        if not os.path.exists(html_file_path):
            return

        clean = [
            {k: _sanitize(v) for k, v in rec.items()}
            for rec in records
        ]
        json_str = json.dumps(clean, indent=2, ensure_ascii=False)

        with open(html_file_path, "r", encoding="utf-8") as fh:
            content = fh.read()

        # Try exact marker approach first
        start_marker = "const reportData = "
        end_marker   = ";\n\n    // State Variables"
        start_idx = content.find(start_marker)
        end_idx   = content.find(end_marker)

        if start_idx != -1 and end_idx != -1:
            new_content = (
                content[:start_idx]
                + start_marker
                + json_str
                + content[end_idx:]
            )
        else:
            import re
            new_content = re.sub(
                r"const\s+reportData\s*=\s*\[[\s\S]*?\];",
                f"const reportData = {json_str};",
                content,
            )

        with open(html_file_path, "w", encoding="utf-8") as fh:
            fh.write(new_content)
